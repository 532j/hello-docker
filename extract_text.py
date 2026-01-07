import os
from openpyxl import load_workbook
from docx import Document

# テキスト保存用のフォルダ
os.makedirs("extracted_texts", exist_ok=True)

def extract():
    files = os.listdir("samples")
    
    for filename in files:
        filepath = os.path.join("samples", filename)
        text_content = ""
        
        # --- エクセルの場合 ---
        if filename.endswith(".xlsx"):
            wb = load_workbook(filepath)
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    # 空白ではないセルを文字列として結合
                    text_content += " ".join([str(cell) for cell in row if cell is not None]) + "\n"
        
        # --- ワードの場合 ---
        elif filename.endswith(".docx"):
            doc = Document(filepath)
            text_content = "\n".join([para.text for para in doc.paragraphs])
        
        # --- テキストファイルとして保存 ---
        if text_content:
            output_name = f"{os.path.splitext(filename)[0]}.txt"
            with open(os.path.join("extracted_texts", output_name), "w", encoding="utf-8") as f:
                f.write(text_content)
            print(f"📄 抽出完了: {filename} -> {output_name}")

if __name__ == "__main__":
    extract()